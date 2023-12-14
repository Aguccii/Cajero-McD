import os
import time
import funciones
import openpyxl as op
from openpyxl.styles import Alignment, alignment


ComboS = 650
ComboD = 700
ComboT = 800
Flurby = 250

wb = op.Workbook()

if not os.path.exists('ejemplo.xlsx'):
	input("dos aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
	ws1 = wb.active



	Cel1 = ws1["A1"] = "Cliente"
	Cel1 = Alignment(horizontal='center')

	Cel2 = ws1["B1"] = "Fecha"
	Cel2 = Alignment(horizontal='center')
	Cel2 = ws1.column_dimensions['B'].width = 25

	Cel3 = ws1["C1"] = "Combo S"
	Cel3 = Alignment(horizontal='center')

	Cel4 = ws1["D1"] = "Combo D"
	Cel4 = Alignment(horizontal='center')

	Cel5 = ws1["E1"] = "Combo T"
	Cel5 = Alignment(horizontal='center')

	Cel6 = ws1["F1"] = "Flurby"
	Cel6 = Alignment(horizontal='center')

	Cel7 = ws1["G1"] = "Total"
	Cel7 = Alignment(horizontal='center')

	wb.save(filename="ejemplo.xlsx")
else:
	pass

def iniciar_menu():

	while True: 
		
		
		cliente = input("Ingrese el nombre del cliente: ")
		cliente = funciones.verificar_vacio(cliente)
		cliente = funciones.Verif_Nombre(cliente)
		tiempo = int(time.strftime('%H'))
		Saludo = funciones.Saludo_tiempo(tiempo, cliente.capitalize())


		cant_ComboS = input("Ingrese cantidad Combo S : ")
		cant_ComboS = funciones.verificar_vacio(cant_ComboS)
		cant_ComboS = funciones.convertir(cant_ComboS)
		
		cant_ComboD = input("Ingrese cantidad Combo D : ")
		cant_ComboD = funciones.verificar_vacio(cant_ComboD)
		cant_ComboD = funciones.convertir(cant_ComboD)


		cant_ComboT = input("Ingrese cantidad Combo T : ")
		cant_ComboT = funciones.verificar_vacio(cant_ComboT)
		cant_ComboT = funciones.convertir(cant_ComboT)


		cant_Flurby = input("Ingrese cantidad Flurby : ")
		cant_Flurby = funciones.verificar_vacio(cant_Flurby)
		cant_Flurby = funciones.convertir(cant_Flurby)

		
		Total = ComboS * cant_ComboS + ComboD * cant_ComboD + ComboT * cant_ComboT + Flurby * cant_Flurby
		

		if Total == 0:
			os.system("cls")
			print("No se ordenó nada. Volviendo al menu.")
			time.sleep(1)
			os.system("cls")
			print("No se ordenó nada. Volviendo al menu..")
			time.sleep(1)
			os.system("cls")
			print("No se ordenó nada. Volviendo al menu...")
			time.sleep(1)
			os.system("cls")
			break
		else:
			print("El total es de ", Total, " pesos.")

		Abono = input("Ingrese con cuanto desea abonar : ")
		Abono = funciones.verificar_vacio(Abono)
		Abono = funciones.convertir(Abono)
		Abono = funciones.verif_vuelto(Total, Abono)
		
		Vuelto = Abono - Total
		
		funciones.total_en_caja += (Abono - Vuelto)
		

		print("Su vuelto es de ", Vuelto, " pesos.")
		print("Desea hacer la compra?")
		opcion = input("Introduzca 's' para sí y 'n' para no. >>>")
		opcion = funciones.verificar_vacio(opcion)
		opcion = funciones.Verif_Nombre(opcion)
		opcion = funciones.Verif_SN(opcion.lower())


		if opcion == "s":
			print("Gracias por comprar en Mc Dowell’s!")
			time.sleep(3)
			os.system("cls")
			print("Volviendo al menu.")
			time.sleep(1)
			os.system("cls")
			print("Volviendo al menu..")
			time.sleep(1)
			os.system("cls")
			print("Volviendo al menu...")
			time.sleep(1)
			os.system("cls")

			myFileName = "ejemplo.xlsx"

			wb = op.load_workbook(filename=myFileName)
			ws = wb['Sheet']


			newRowLocation = ws.max_row +4


			ws.append([cliente.capitalize(), time.asctime(), cant_ComboS, cant_ComboD, cant_ComboT, cant_Flurby, Total])
			wb.save(filename=myFileName)
			wb.close()
			break
		else:
			print("Su orden fue cancelada.")
			time.sleep(5)
			os.system("cls")
			print("Volviendo al menu.")
			time.sleep(1)
			os.system("cls")
			print("Volviendo al menu..")
			time.sleep(1)
			os.system("cls")
			print("Volviendo al menu...")
			time.sleep(1)
			os.system("cls")
			break





  


